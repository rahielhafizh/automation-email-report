import pyodbc
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from services.db_connection import get_database_connection


def fetch_dashboard_control_data(conn):
    try:
        cursor = conn.cursor()
        query = "SELECT * FROM [SFI_DWH].[dbo].[Dashboard_Control]"
        cursor.execute(query)
        
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        
        cursor.close()
        return columns, rows
    except pyodbc.Error as e:
        print(f"[ERROR] QUERY FAILED : {e}")
        return None, None


def create_excel_file(columns, rows, filename=None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Dashboard_Control_{timestamp}.xlsx"
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Dashboard_Control"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # WRITE HEADERS
        for col_idx, column_name in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # WRITE DATA ROWS
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # AUTO-ADJUST COLUMN WIDTHS
        for col_idx, column_name in enumerate(columns, start=1):
            max_length = len(str(column_name))
            for row_idx in range(2, len(rows) + 2):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
        
        workbook.save(filename)
        print(f"[SYSTEM] EXCEL FILE CREATED : {filename}")
        return filename
    except Exception as e:
        print(f"[ERROR] EXCEL CREATION FAILED : {e}")
        return None


def export_dashboard_control(output_filename=None):
    conn = get_database_connection()
    
    if conn is None:
        print("[ERROR] DATABASE CONNECTION UNAVAILABLE")
        return False
    
    try:
        columns, rows = fetch_dashboard_control_data(conn)
        
        if columns is None or rows is None:
            return False
        
        if len(rows) == 0:
            print("[WARNING] NO DATA FOUND IN Dashboard_Control TABLE")
            return False
        
        print(f"[SYSTEM] FETCHED {len(rows)} ROWS FROM Dashboard_Control")
        
        result = create_excel_file(columns, rows, output_filename)
        
        return result is not None
    finally:
        conn.close()
        print("[SYSTEM] CONNECTION CLOSED.")


if __name__ == "__main__":
    export_dashboard_control()