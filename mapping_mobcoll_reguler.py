import os
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Callable
from services.config import logger

ASSET_DIR = Path(__file__).parent / "asset"
INPUT_FILE = ASSET_DIR / "Mapping-PIC.xlsx"
OUTPUT_FILE = ASSET_DIR / "Mapping-Result.xlsx"
RE_NON_DIGIT = re.compile(r"[^0-9]")
RE_MULTI_SPACE = re.compile(r"\s{2,}")
RE_AREA_PREFIX = re.compile(r"^[\d\W_]+\s*")
RE_WHITESPACE = re.compile(r"\s+")
BARE_M_PREFIX = re.compile(r"^M$", re.IGNORECASE)

DEGREE_MAP: dict[str, str] = {
    "AMD": "A.Md.",
    "AMDKOM": "A.Md.Kom.",
    "AMDAK": "A.Md.Ak.",
    "AMDM": "A.Md.M.",
    "AMDAB": "A.Md.A.B.",
    "AMDPJK": "A.Md.Pjk.",
    "AMDPAR": "A.Md.Par.",
    "STR": "S.Tr.",
    "STRKOM": "S.Tr.Kom.",
    "STRAK": "S.Tr.Ak.",
    "STRAB": "S.Tr.A.B.",
    "STRM": "S.Tr.M.",
    "SE": "S.E.",
    "SAK": "S.Ak.",
    "SM": "S.M.",
    "SAB": "S.A.B.",
    "SAP": "S.A.P.",
    "SKOM": "S.Kom.",
    "SI": "S.I.",
    "SH": "S.H.",
    "SIKOM": "S.I.Kom.",
    "ST": "S.T.",
    "SSOS": "S.Sos.",
    "SPSI": "S.Psi.",
    "SMAT": "S.Mat.",
    "STAT": "S.Stat.",
    "SSI": "S.Si.",
    "SPD": "S.Pd.",
    "SP": "S.P.",
    "SS": "S.S.",
    "SHUM": "S.Hum.",
    "SDES": "S.Ds.",
    "SDS": "S.Ds.",
}

DEGREE_KEYS_UPPER: frozenset[str] = frozenset(DEGREE_MAP.keys())

FRONT_TITLES_MAP: dict[str, str] = {
    "DRS": "Drs.",
    "DRA": "Dra.",
    "IR": "Ir.",
    "DR": "Dr.",
    "PROF": "Prof.",
    "DRG": "drg.",
    "DRH": "drh.",
    "NS": "Ns.",
    "BD": "Bd.",
    "APT": "apt.",
    "H": "H.",
}

PREFIX_ABBREV: dict[str, str] = {
    "MUHAMMAD": "M.",
    "MUHAMAD": "M.",
    "MOCHAMMAD": "M.",
    "MOHAMMAD": "M.",
    "MOCHAMAD": "M.",
    "MOHAMAD": "M.",
    "MOH": "M.",
}

BRANCH_COMPOUNDS: frozenset[str] = frozenset(
    {
        "BANDAR LAMPUNG",
        "DEWI SARTIKA",
        "PANGKAL PINANG",
    }
)

BRANCH_TOKENS: frozenset[str] = frozenset(
    {
        "BALIKPAPAN",
        "BANDUNG",
        "BANJARMASIN",
        "BARABAI",
        "BATAM",
        "BEKASI",
        "BOGOR",
        "CIREBON",
        "DENPASAR",
        "DEPOK",
        "GORONTALO",
        "GRESIK",
        "JAMBI",
        "KARAWANG",
        "KEDIRI",
        "KEDOYA",
        "KENDARI",
        "KUDUS",
        "KUPANG",
        "MAKASSAR",
        "MALANG",
        "MANADO",
        "MATARAM",
        "MEDAN",
        "PADANG",
        "PALANGKARAYA",
        "PALEMBANG",
        "PALU",
        "PEKANBARU",
        "PONTIANAK",
        "PURWOKERTO",
        "SAMARINDA",
        "SAMPIT",
        "SEMARANG",
        "SERANG",
        "SOLO",
        "SUNTER",
        "SURABAYA",
        "TANGERANG",
        "TEGAL",
        "TERNATE",
        "YOGYAKARTA",
        "BPN",
        "BDL",
        "BDG",
        "BJM",
        "BRB",
        "BTM",
        "BKS",
        "BGR",
        "CRB",
        "DPS",
        "DPK",
        "GTO",
        "GRK",
        "JMB",
        "KRW",
        "KDR",
        "KDY",
        "KDI",
        "KDS",
        "KPG",
        "MKS",
        "MLG",
        "MND",
        "MTR",
        "MDN",
        "PDG",
        "PLK",
        "PLG",
        "PLW",
        "PKP",
        "PBR",
        "PTK",
        "PWT",
        "SMD",
        "SMP",
        "SMG",
        "SRG",
        "SLO",
        "STR",
        "SBY",
        "TGR",
        "TGL",
        "TTE",
        "YGY",
        "CABANG",
        "AREA",
        "REGION",
        "KOTA",
        "KABUPATEN",
    }
)

BALINESE_COMPOUND: frozenset[str] = frozenset(
    {
        "I MADE",
        "I WAYAN",
        "I KADEK",
        "I GEDE",
        "I PUTU",
        "I KOMANG",
        "I NGURAH",
        "I KETUT",
        "NI LUH",
        "NI MADE",
        "NI WAYAN",
        "NI KADEK",
        "NI KETUT",
        "NI KOMANG",
        "NI PUTU",
        "IDA BAGUS",
        "IDA AYU",
        "I NYOMAN",
        "NI NYOMAN",
        "I GUSTI",
        "I GUSTI AYU",
        "ANAK AGUNG",
        "COKORDA",
        "TJOKORDA",
        "GUSTI AYU",
        "GUSTI NGURAH",
        "DEWA AYU",
        "DEWA GEDE",
        "ANAK AGUNG ISTRI",
        "ANAK AGUNG GEDE",
    }
)

BALINESE_SINGLE: frozenset[str] = frozenset(
    {"I", "NI", "IDA", "ANAK", "DEWA", "DESAK", "GUSTI"}
)
PROTECTED_SINGLES: frozenset[str] = frozenset({"I", "M."})


def get_balinese_prefix_len(upper_tokens: list[str]) -> int:
    if len(upper_tokens) >= 3 and " ".join(upper_tokens[:3]) in BALINESE_COMPOUND:
        return 3
    if len(upper_tokens) >= 2 and " ".join(upper_tokens[:2]) in BALINESE_COMPOUND:
        return 2
    if upper_tokens and upper_tokens[0] in BALINESE_SINGLE:
        return 1
    return 0


def format_token(token: str) -> str:
    if "." in token:
        return token
    upper = token.upper()
    if len(token) == 1 and token.isalpha():
        return upper if upper in PROTECTED_SINGLES else upper + "."
    return token.title()


def clean_name(value) -> str:
    raw = str(value).strip() if pd.notna(value) else ""
    if not raw:
        return ""

    raw = RE_MULTI_SPACE.sub(" ", raw).strip().rstrip(".,")
    parts = [p.strip() for p in raw.split(",")]
    core_raw, comma_degrees_raw = parts[0], parts[1:]

    tokens_u = core_raw.upper().split()
    if not tokens_u:
        return ""

    front_titles = []
    while tokens_u and tokens_u[0].replace(".", "") in FRONT_TITLES_MAP:
        front_titles.append(FRONT_TITLES_MAP[tokens_u.pop(0).replace(".", "")])

    while tokens_u:
        popped = False
        if len(tokens_u) >= 2:
            last_two = f"{tokens_u[-2]} {tokens_u[-1]}"
            if last_two in BRANCH_COMPOUNDS:
                del tokens_u[-2:]
                popped = True
                continue
        if tokens_u[-1] in BRANCH_TOKENS:
            tokens_u.pop()
            popped = True
            continue
        break

    inline_degrees = []
    while tokens_u and tokens_u[-1].replace(".", "") in DEGREE_KEYS_UPPER:
        inline_degrees.insert(0, DEGREE_MAP[tokens_u.pop().replace(".", "")])

    if not tokens_u:
        return " ".join(front_titles) if front_titles else raw

    if tokens_u[0] in PREFIX_ABBREV:
        tokens_u[0] = PREFIX_ABBREV[tokens_u[0]]
    elif BARE_M_PREFIX.match(tokens_u[0]) and len(tokens_u) > 1:
        tokens_u[0] = "M."

    keep_full = (
        get_balinese_prefix_len(tokens_u) + 1
        if get_balinese_prefix_len(tokens_u) > 0
        else 2
    )

    if len(tokens_u) <= keep_full:
        core_name = " ".join(format_token(t) for t in tokens_u)
    else:
        full_part = " ".join(format_token(t) for t in tokens_u[:keep_full])
        letters = [t[0].upper() for t in tokens_u[keep_full:] if t and t[0].isalpha()]
        initials = ".".join(letters) + "." if letters else ""
        core_name = f"{full_part} {initials}".strip()

    final_name = f"{' '.join(front_titles)} {core_name}".strip()

    all_degrees = inline_degrees.copy()
    for cd in comma_degrees_raw:
        clean_cd = cd.upper().replace(".", "").replace(" ", "")
        if clean_cd in DEGREE_MAP:
            all_degrees.append(DEGREE_MAP[clean_cd])
            continue
        sub_parts = cd.split()
        if len(sub_parts) > 1:
            sub_clean = "".join(sub_parts).upper().replace(".", "")
            if sub_clean in DEGREE_MAP:
                all_degrees.append(DEGREE_MAP[sub_clean])
                continue
        all_degrees.append(cd.title())

    result_str = (
        f"{final_name}, {', '.join(all_degrees)}" if all_degrees else final_name
    )
    return result_str.rstrip(".")


def clean_branch_id(value) -> str:
    digits = RE_NON_DIGIT.sub("", str(value) if pd.notna(value) else "")
    return digits.zfill(4) if digits else ""


def clean_area(value) -> str:
    raw = str(value).strip() if pd.notna(value) else ""
    collapsed = RE_MULTI_SPACE.sub(" ", RE_AREA_PREFIX.sub("", raw)).strip().upper()
    return "UNKNOWN_AREA" if raw and not collapsed else collapsed


def clean_position(value) -> str:
    return RE_WHITESPACE.sub("", str(value).strip() if pd.notna(value) else "").upper()


def clean_phone_number(value) -> str:
    raw = str(value).strip() if pd.notna(value) else ""
    if not raw:
        return ""
    digits = RE_NON_DIGIT.sub("", raw)
    if digits.startswith("62"):
        return "0" + digits[2:]
    if digits.startswith("8"):
        return "0" + digits
    return digits


ALT_ROW_FILL = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
CENTERED_COLUMNS: frozenset[int] = frozenset({1, 2, 3, 6, 10, 14})
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMN_CLEANERS: dict[int, Callable[[object], str]] = {
    2: clean_branch_id,
    3: clean_area,
    6: clean_name,
    7: clean_position,
    10: clean_name,
}

SORT_PRIORITY_INDICES: tuple[int, ...] = (2, 6, 10)


def apply_worksheet_style(ws, col_count: int, col_max_lengths: dict[int, int]) -> None:
    for col in range(1, col_count + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.font, header_cell.fill = HEADER_FONT, HEADER_FILL
        header_cell.alignment, header_cell.border = CENTER, THIN_BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, cell in enumerate(row, start=1):
            cell.font, cell.fill, cell.border = BODY_FONT, row_fill, THIN_BORDER
            cell.alignment = CENTER if col_idx in CENTERED_COLUMNS else LEFT

    for col in range(1, col_count + 1):
        col_letter = get_column_letter(col)
        max_len = col_max_lengths.get(col - 1, 10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"


def load_source(file_path: Path) -> pd.DataFrame:
    if not file_path.exists():
        raise FileNotFoundError(f"Source file not found: {file_path}")
    return pd.read_excel(file_path, dtype=str)


def sort_by_priority(df: pd.DataFrame) -> pd.DataFrame:
    columns = df.columns
    sort_cols = [columns[i] for i in SORT_PRIORITY_INDICES if len(columns) > i]
    if not sort_cols:
        return df
    return df.sort_values(
        by=sort_cols,
        key=lambda col: col.str.upper(),
        kind="mergesort",
        ignore_index=True,
    )


def transform(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    columns = result.columns

    for index, cleaner in COLUMN_CLEANERS.items():
        if len(columns) > index:
            result[columns[index]] = result[columns[index]].map(cleaner)

    phone_index = next(
        (
            i
            for i, col in enumerate(columns)
            if "HP" in str(col).upper() or "TELEPON" in str(col).upper()
        ),
        None,
    )
    if phone_index is not None:
        result[columns[phone_index]] = result[columns[phone_index]].map(
            clean_phone_number
        )

    return sort_by_priority(result)


def write_output(df: pd.DataFrame, file_path: Path) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)

    col_max_lengths = {
        i: max(
            df[col].astype(str).map(len).max() if not df.empty else 10, len(str(col))
        )
        for i, col in enumerate(df.columns)
    }

    df.to_excel(file_path, index=False, sheet_name="Result")

    wb = load_workbook(file_path)
    apply_worksheet_style(wb.active, len(df.columns), col_max_lengths)
    wb.save(file_path)


def run() -> None:
    os.system("cls" if os.name == "nt" else "clear")
    logger.info("[INFO] MEMULAI PROSES MAPPING DATA PIC")

    try:
        logger.info("[INFO] MENARIK DATA DARI SUMBER")
        df_raw = load_source(INPUT_FILE)
        logger.info(f"[INFO] DATA BERHASIL DITARIK ({len(df_raw)} BARIS).")

        logger.info("[INFO] MEMULAI TRANSFORMASI DATA")
        df_clean = transform(df_raw)
        logger.info("[INFO] TRANSFORMASI DATA SELESAI")

        logger.info("[INFO] MENYIMPAN HASIL KE FORMAT EXCEL")
        write_output(df_clean, OUTPUT_FILE)
        logger.info(f"[INFO] HASIL TERSIMPAN PADA : {OUTPUT_FILE.resolve()}")

    except FileNotFoundError as e:
        logger.error(f"[ERROR] FILE TIDAK DITEMUKAN: {e}")
    except Exception as e:
        logger.error(f"[ERROR] TERJADI KESALAHAN : {e}")


if __name__ == "__main__":
    run()